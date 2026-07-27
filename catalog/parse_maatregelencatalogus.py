"""
Parser: Nij Begun Maatregelencatalogus (xlsx) -> catalog.json (schoon, machine-leesbaar).

- Prijslijst uit sheet 'Prijsonderbouwing': code, omschrijving, eenheid, prijs/ehd (excl), incl. btw.
- Rc-per-dikte uit sheet 'Tabellen' (best-effort): per materiaalvariant dikte->Rc.

NB: dit is een offline fallback. Zodra je de Nij Begun catalogus-API-key hebt
(leveranciers@nijbegun.nl) gebruiken we die als bron van waarheid; deze parser blijft
nuttig als offline cache en voor de Rc-tabellen.

Gebruik:
    python3 catalog/parse_maatregelencatalogus.py "../../Maatregelencatalogus-...xlsx"
"""
import sys, os, re, json, datetime
import openpyxl

ONDERDEEL = {"V1": "A Gevel", "V2": "B Glas en kozijnen", "V3": "C Vloeren",
             "V4": "D Daken", "V5": "E Ventilatie"}
CODE_RE = re.compile(r'^V\d')
SPEC_CODE_RE = re.compile(r'^V\d+-')          # echte maatregel (bv V1-1-A1, V3-1)
RC_RE = re.compile(r'(\d+(?:[.,]\d+)?)\s*(mm|cm)\s*Rc\s*([\d.,]+)', re.I)

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(float(v), 4)
    s = str(v).strip().replace(',', '.')
    try: return round(float(s), 4)
    except ValueError: return None

def onderdeel_for(code):
    m = re.match(r'^(V\d)', code)
    return ONDERDEEL.get(m.group(1), "") if m else ""

def parse_prijslijst(ws):
    maatregelen, seen = [], set()
    current_level = ""
    for r in ws.iter_rows(min_row=4, values_only=True):
        code = (str(r[1]).strip() if len(r) > 1 and r[1] is not None else "")
        if not code or not CODE_RE.match(code):
            continue
        oms = (str(r[3]).strip() if len(r) > 3 and r[3] else "")
        # level-header (bv 'V1 ' + 'Level 1 - GEVEL')
        if not SPEC_CODE_RE.match(code):
            current_level = oms or current_level
            continue
        prijs_excl = num(r[14]) if len(r) > 14 else None
        prijs_incl = num(r[23]) if len(r) > 23 else None
        eenheid = (str(r[5]).strip() if len(r) > 5 and r[5] else "")
        key = (code, oms)
        if key in seen:
            # vul incl. btw aan als die er nu wel is
            for m in maatregelen:
                if (m["code"], m["omschrijving"]) == key and m["prijs_per_eenheid_incl_btw"] is None:
                    m["prijs_per_eenheid_incl_btw"] = prijs_incl
            continue
        if prijs_excl is None and prijs_incl is None:
            continue
        seen.add(key)
        maatregelen.append({
            "code": code, "onderdeel": onderdeel_for(code), "level": current_level,
            "omschrijving": oms, "eenheid": eenheid,
            "prijs_per_eenheid_excl": prijs_excl,
            "prijs_per_eenheid_incl_btw": prijs_incl,
        })
    return maatregelen

def parse_rc_tabellen(ws):
    """Best-effort: zoek '... meer-/minderprijs per mm/cm'-headers en lees dikte->Rc eronder."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    tabellen = []
    for ri, r in enumerate(rows):
        for ci, cell in enumerate(r or []):
            if not cell or "meer-/minderprijs per" not in str(cell).lower():
                continue
            mat = re.sub(r'^[╚\s]+', '', str(cell)).split("meer-/minderprijs")[0].strip()
            # component-naam: zoek omhoog in dezelfde kolom naar een korte tekst
            component = ""
            for up in range(ri - 1, max(ri - 4, -1), -1):
                c = rows[up][ci] if ci < len(rows[up]) else None
                if c and "prijs" not in str(c).lower():
                    component = str(c).strip(); break
            stappen = []
            for rr in range(ri + 1, min(ri + 12, len(rows))):
                row = rows[rr]
                delta = row[ci] if ci < len(row) else None
                # de Rc-string staat 1 of 2 kolommen rechts
                rc_cell = None
                for off in (1, 2):
                    if ci + off < len(row) and row[ci + off] and RC_RE.search(str(row[ci + off])):
                        rc_cell = str(row[ci + off]); break
                if rc_cell is None:
                    if delta is None: break
                    continue
                m = RC_RE.search(rc_cell)
                stappen.append({
                    "prijs_delta": num(delta),
                    "dikte": num(m.group(1)), "dikte_eenheid": m.group(2).lower(),
                    "rc": num(m.group(3)),
                })
            if stappen:
                tabellen.append({"component": component, "materiaal": mat, "stappen": stappen})
    return tabellen

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else \
        "../../Maatregelencatalogus-aannemer-aangepast-V3_Q2_03062026.xlsx"
    here = os.path.dirname(os.path.abspath(__file__))
    xlsx = xlsx if os.path.isabs(xlsx) else os.path.join(here, xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    maatregelen = parse_prijslijst(wb["Prijsonderbouwing"])
    rc_tab = parse_rc_tabellen(wb["Tabellen"])
    wb.close()
    # versie uit de bestandsnaam: 'V3_Q2_03062026' (oud) of 'Q3_2026_21072026' (vanaf Q3-2026)
    versie = (re.search(r'(V\d+_Q\d+_\d+)', os.path.basename(xlsx))
              or re.search(r'(Q\d+_\d{4}_\d+)', os.path.basename(xlsx)))
    catalog = {
        "bron": os.path.basename(xlsx),
        "versie": versie.group(1) if versie else "onbekend",
        "gegenereerd_op": datetime.date.today().isoformat(),
        "aantal_maatregelen": len(maatregelen),
        "maatregelen": maatregelen,
        "rc_tabellen": rc_tab,
    }
    out = os.path.join(here, "catalog.json")
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(catalog, fh, ensure_ascii=False, indent=2)
    print(f"OK: {out}")
    print(f"  maatregelen: {len(maatregelen)} | rc-tabellen: {len(rc_tab)}")
    per = {}
    for m in maatregelen:
        per[m["onderdeel"]] = per.get(m["onderdeel"], 0) + 1
    for k, v in sorted(per.items()):
        print(f"    {k or '(geen onderdeel)'}: {v}")
    print("  voorbeelden:")
    for m in maatregelen[:4]:
        print(f"    {m['code']:10s} €{m['prijs_per_eenheid_excl']}/{m['eenheid']}  {m['omschrijving'][:55]}")
    if rc_tab:
        t = rc_tab[0]
        print(f"  rc-tabel voorbeeld [{t['component']} / {t['materiaal']}]: "
              + ", ".join(f"{s['dikte']}{s['dikte_eenheid']}→Rc{s['rc']}" for s in t['stappen'][:4]))

if __name__ == "__main__":
    main()
